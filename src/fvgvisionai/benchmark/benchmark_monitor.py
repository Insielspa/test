import logging
import os.path
import threading
import traceback
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from numpy import ndarray

import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from io import BytesIO
from openpyxl.chart import Reference, LineChart
from openpyxl.chart.series import SeriesLabel
from openpyxl.worksheet.worksheet import Worksheet

from fvgvisionai.common.app_timer import AppTimer
from fvgvisionai.common.atomic_boolean import AtomicBoolean
from fvgvisionai.config.app_settings import AppSettings
from fvgvisionai.processor.data_aggregator import DataAggregator

X_AXIS_TITLE = "Time (s)"

DATE_FORMAT = "%Y-%m-%d %H:%M:%S.%f"

DEFAULT_AGGREGATION_TIME_S = 10.0


def _get_now_as_string():
    return datetime.today().strftime(DATE_FORMAT)[:-3]


def _get_datetime_as_string(date: datetime):
    return date.strftime(DATE_FORMAT)[:-3]


class BenchmarkMonitor:
    def __init__(self, enabled: bool, app_settings: AppSettings, measures_aggregation_time_ms: int,
                 exit_signal: threading.Event):
        self._enabled = enabled
        self._lock = threading.Lock()
        self._processor_timer = AppTimer()
        self._logger = logging.getLogger(__name__)
        self._ready = AtomicBoolean()
        self._message_counter = 0
        self._ready.set(True)
        self._app_settings = app_settings
        self._measures_aggregation_time_ms = measures_aggregation_time_ms
        self._acquire_frame = False
        self._exit_signal = exit_signal
        self._executor = ThreadPoolExecutor(thread_name_prefix="benchmark")

        self._interval_timer = AppTimer()
        self._benchmark_timer = AppTimer()

        # Creare un nuovo foglio Excel
        if os.path.isfile(self._app_settings.benchmark_results_file_name):
            self.workbook = openpyxl.load_workbook(self._app_settings.benchmark_results_file_name, read_only=False)
        else:
            self.workbook = openpyxl.Workbook()

        if self._app_settings.model_use_tensort:
            model_type = app_settings.model_precision.value['suffix'] + " " \
                         + app_settings.model_resolution.value['suffix']
        else:
            model_type = "pytorch"

        self.bench_name = app_settings.model_id + " " + model_type
        self.sheet_name = f"b {self.bench_name}"

        if self.sheet_name in self.workbook.get_sheet_names():
            self.current_sheet = self.workbook.get_sheet_by_name(self.sheet_name)
            self.workbook.remove(self.current_sheet)

        self.current_sheet = self.workbook.create_sheet(self.sheet_name)

        headers1 = [
            "time",
            "model_fps",
            "model_time_ms",
            "max_people",
            "min_people",
            "avg_people",
            "avg_bikes",
            "avg_cars",
            "max_people_in_zone",
            "min_people_in_zone",
            "avg_people_in_zone",
            "max_time_in_zone",
            "min_time_in_zone",
            "avg_time_in_zone",
            "sum_entrances",
            "sum_exits"
        ]
        self.current_sheet.append(headers1)

    def start(self):
        self._interval_timer.start()
        self._benchmark_timer.start()

    def close(self):

        for sheet_name in self.workbook.sheetnames:
            if not sheet_name.startswith("b "):
                self.workbook.remove(self.workbook.get_sheet_by_name(sheet_name))

        self.generate_graph(sheet_title="g fps", graph_title="Model Per Second", x_axis_title=X_AXIS_TITLE,
                            data_column=2, y_axis_title="Model Frame Per Second")
        self.generate_graph(sheet_title="g process time", graph_title="Inference Time", x_axis_title=X_AXIS_TITLE,
                            data_column=3, y_axis_title="Time for Model")
        self.generate_graph(sheet_title="g avg people", graph_title="Average people", x_axis_title=X_AXIS_TITLE,
                            data_column=6, y_axis_title="AVG People")
        self.generate_graph(sheet_title="g avg car", graph_title="Average car", x_axis_title=X_AXIS_TITLE,
                            data_column=8, y_axis_title="AVG Car")

        # Salvare il foglio Excel
        self.workbook.save(self._app_settings.benchmark_results_file_name)

        # Chiudere il foglio Excel
        self.workbook.close()

    def _build_message_payload(self,
                               data: DataAggregator) -> list:
        agg_frame_info_list = [
            round(self._benchmark_timer.elapsed_time / 1_000),
            round(1_000 / max(data.time_frame_processing_average, 1), 2),
            data.time_frame_processing_average,
            data.people_counter.max,
            data.people_counter.min,
            data.people_counter.average,
            data.bikes_counter.average,
            data.cars_counter.average,
            data.people_in_zone_counter.max,
            data.people_in_zone_counter.min,
            data.people_in_zone_counter.average,
            data.max_time_in_zone,
            data.max_time_in_zone,
            data.avg_time_in_zone,
            data.door_people_entered,
            data.door_people_leaved
        ]
        return agg_frame_info_list

    def generate_graph(self, sheet_title: str, graph_title: str, x_axis_title: str, y_axis_title: str,
                       data_column: int):
        # Creare foglio 3 per il grafico
        chart_sheet = self.workbook.create_sheet(title=sheet_title)
        chart = LineChart()

        index = 0
        for sheet_name in self.workbook.sheetnames:
            if sheet_name.startswith("b "):
                current_sheet: Worksheet = self.workbook.get_sheet_by_name(sheet_name)
                values = Reference(current_sheet, min_col=data_column, min_row=1, max_col=data_column,
                                   max_row=current_sheet.max_row)
                chart.add_data(values, titles_from_data=True)
                label = SeriesLabel(v=sheet_name.replace("b ", ""))
                chart.series[index].title = label
                index += 1
        # https://stackoverflow.com/questions/48661658/making-graph-openpyxl-categories-not-selected
        categories = Reference(self.current_sheet, min_col=1, max_col=1, min_row=2, max_row=self.current_sheet.max_row)
        chart.set_categories(categories)

        chart.title = graph_title
        chart.x_axis.title = x_axis_title
        chart.y_axis.title = y_axis_title
        # Inserire il grafico nella foglio del grafico
        chart_sheet.add_chart(chart, "A1")

    def measure_performance(self, frame_index: int, source_frame: ndarray, data_aggregator: DataAggregator) -> bool:
        global_elapsed_time_ms = self._benchmark_timer.elapsed_time
        elapsed_time_from_last_benchmark_ms = self._interval_timer.elapsed_time

        if global_elapsed_time_ms < self._app_settings.benchmark_warmup_time_ms:
            return False
        elif elapsed_time_from_last_benchmark_ms >= self._measures_aggregation_time_ms:
            data = data_aggregator.copy()
            self._executor.submit(self._registry_data_for_benchmark, frame_index, source_frame, data)
            self._interval_timer.start()
            return True
        elif (global_elapsed_time_ms > self._app_settings.benchmark_duration_time_ms +
              self._app_settings.benchmark_warmup_time_ms):
            self._exit_signal.set()
            return False

        return False

    def _registry_data_for_benchmark(self, frame_index: int, source_frame: ndarray, data: DataAggregator):
        with self._lock:
            try:
                self._ready.set(False)

                message = self._build_message_payload(data)
                self._message_counter = (self._message_counter + 1) % 1_000_000

                self.current_sheet.append(message)

                self._logger.info(
                    f"Benchmark #{self._message_counter} data with  {data.time_frame_processing_average:5.1f} ms")
                self._logger.debug(f"message payload: {message}")

                if not self._acquire_frame and frame_index > 100:
                    self._acquire_frame = True

                    # Converte l'array NumPy in un'immagine PIL
                    pil_image = PILImage.fromarray(source_frame)

                    # Converte l'immagine PIL in bytes
                    image_bytes_io = BytesIO()
                    pil_image.save(image_bytes_io, format='PNG')

                    # Aggiungi l'immagine al foglio di lavoro senza salvarla su disco
                    img = Image(image_bytes_io)

                    self.current_sheet.add_image(img, "R1")
            except GeneratorExit:
                # This is raised when the client disconnects
                logging.error('Client disconnected')
            except Exception as e:
                # Gestisci l'eccezione qui
                logging.error(f"Si è verificata un'eccezione: {e}")
                traceback.print_exc()
            finally:
                self._ready.set(True)
